import openpyxl
import json
import sys
from pathlib import Path

print('=' * 60)
print('Excel 轉 JSON 工具 (Python 版本)')
print('=' * 60)

# 設定檔案路徑
input_file = Path('H:/Antigravity make/game-database/Digimon wiki.xlsx')
output_file = Path('H:/Antigravity make/game-database-github/data/sample-data.json')

# 檢查輸入檔案
if not input_file.exists():
    print(f'❌ 錯誤：找不到檔案 {input_file}')
    sys.exit(1)

try:
    print(f'📖 正在讀取 Excel 檔案...')
    print(f'   檔案位置：{input_file}')
    
    # 載入 Excel 檔案
    wb = openpyxl.load_workbook(input_file, data_only=True)
    
    print(f'✅ 成功讀取 Excel 檔案')
    print(f'   工作表數量：{len(wb.sheetnames)}')
    print(f'   工作表列表：{", ".join(wb.sheetnames)}')
    
    # 轉換所有工作表
    data = {}
    total_rows = 0
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = []
        
        # 讀取所有列
        for row in sheet.iter_rows(values_only=True):
            # 將 None 轉換為空字串，保持資料一致性
            row_data = [cell if cell is not None else '' for cell in row]
            sheet_data.append(row_data)
        
        data[sheet_name] = sheet_data
        total_rows += len(sheet_data)
        print(f'   - {sheet_name}: {len(sheet_data)} 列')
    
    print('')
    print('💾 正在儲存 JSON 檔案...')
    print(f'   輸出位置：{output_file}')
    
    # 確保輸出目錄存在
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # 儲存為 JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    # 取得檔案大小
    file_size_kb = output_file.stat().st_size / 1024
    
    print('')
    print('=' * 60)
    print('✅ 轉換完成！')
    print('=' * 60)
    print(f'📊 統計資訊：')
    print(f'   - 工作表數量：{len(wb.sheetnames)}')
    print(f'   - 總列數：{total_rows}')
    print(f'   - 檔案大小：{file_size_kb:.2f} KB')
    print('')
    print('📝 下一步：')
    print('   1. 開啟 index.html')
    print('   2. 網頁會自動載入 sample-data.json')
    print('   3. 開始使用！')
    print('')
    
except Exception as e:
    print(f'❌ 轉換失敗：{str(e)}')
    print('')
    print('可能的原因：')
    print('   - Excel 檔案格式不正確')
    print('   - 檔案已被其他程式開啟')
    print('   - 權限不足')
    print('   - 缺少 openpyxl 套件（請執行：pip install openpyxl）')
    sys.exit(1)
